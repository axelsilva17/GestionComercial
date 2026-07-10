from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random

wb = Workbook()
ws = wb.active
ws.title = "Productos"

# Header style
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2563EB")
hdr_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="94A3B8"),
    right=Side(style="thin", color="94A3B8"),
    top=Side(style="thin", color="94A3B8"),
    bottom=Side(style="thin", color="94A3B8"),
)

headers = ["Nombre", "CodigoBarra", "PrecioVenta", "PrecioCosto", "StockActual", "StockMinimo", "Categoria", "UnidadMedida"]
cats = ["Electrónicos", "Hogar", "Limpieza", "Librería", "Indumentaria", "Cocina", "Jardín", "Bazar", "Perfumería", "Ferretería"]
unidades = ["Unidad", "KG", "Lt", "Pack", "Caja"]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 12

rows = 2500
for i in range(rows):
    num = i + 1
    cat = cats[i % len(cats)]
    ws.cell(row=i+2, column=1, value=f"Producto de Prueba #{num}")
    ws.cell(row=i+2, column=2, value=f"BARR{num:06d}" if num <= 2000 else "")
    ws.cell(row=i+2, column=3, value=random.randint(500, 50000))
    ws.cell(row=i+2, column=4, value=random.randint(200, 30000))
    ws.cell(row=i+2, column=5, value=random.randint(0, 200))
    ws.cell(row=i+2, column=6, value=10)
    ws.cell(row=i+2, column=7, value=cat)
    ws.cell(row=i+2, column=8, value=unidades[i % len(unidades)])

out = r"C:\Users\Usuario\Desktop\GestionComercial\test_import_2500_productos.xlsx"
wb.save(out)
print(f"Generado: {out} - {rows} productos")
